from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font, colors
from config import MANUFACTURERS_FILE_PATH


def xls_create_cheet(workbook, cheet_name='Батарейки'):
    worksheet = workbook.create_sheet(cheet_name, 0)
    return worksheet


def xls_set_cell_value(worksheet, value, row_num, coll_num, bolt=False):
    cell = worksheet.cell(row=row_num, column=coll_num)
    cell.value = value
    xls_set_cell_appearance(cell, bolt=bolt)
    return cell


def xls_add_horizontal_titles(worksheet, titles, bolt=True):
    for column_num, title in enumerate(titles, 2):
        row_num = 1
        cell = xls_set_cell_value(worksheet, title, row_num, column_num)
        xls_set_cell_appearance(cell, bolt=bolt)
    return worksheet


def xls_add_vertical_titles(worksheet, titles, bolt=True):
    for row_num, title in enumerate(titles, 2):
        column_num = 1
        cell = xls_set_cell_value(worksheet, title, row_num, column_num)
        xls_set_cell_appearance(cell, bolt=bolt)
    return worksheet


def xls_save_workbook(workbook, file_name):
    workbook.save(file_name)
    return workbook


def xls_expands_all_columns_width(worksheet):
    def as_text(value):
        if value is None:
            return ""
        return str(value)
    for column_cells in worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length
    return worksheet


def xls_set_cell_appearance(table_cell,
                            border=True,
                            bolt=True,
                            horizontal_alignment='center',
                            u=None,
                            color=colors.BLACK):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    table_cell.font = Font(bold=bolt, u=u, color=color)
    table_cell.alignment = Alignment(horizontal=horizontal_alignment)
    if border:
        table_cell.border = thin_border


def read_file(file_name):
    with open(file_name, 'r') as file:
        lines = file.readlines()
    return lines


battery_log_test_params = {0: {'elapsed_time': 132.61222222222221,
                               'capacity_battery': 925.7659233333333,
                               'mv_average': 2603.9558560857363},
                           1: {'elapsed_time': 132.03611111111113,
                               'capacity_battery': 921.7440916666668,
                               'mv_average': 2634.8081752947205},
                           2: {'elapsed_time': 138.71083333333334,
                               'capacity_battery': 968.3403275000001,
                               'mv_average': 2648.001707525308}
                           }


def creator(battery_log_data):
    vertical_titles = ['Время разряда, ч', 'Ёмкость mA*h', 'Ёмкость mW*h']
    titles_lines = read_file(MANUFACTURERS_FILE_PATH)
    workbook = Workbook()
    worksheet = xls_create_cheet(workbook)
    xls_add_horizontal_titles(worksheet, titles_lines)
    xls_add_vertical_titles(worksheet, vertical_titles)

    for battery_manufacturer_num in battery_log_data.keys():
        column_num = battery_manufacturer_num + 2
        battery_stats = battery_log_data[battery_manufacturer_num]
        discharge_time = battery_stats['elapsed_time']
        mah_result = battery_stats['mv_average']
        mwh_result = battery_stats['capacity_battery']
        xls_set_cell_value(worksheet,
                           discharge_time,
                           row_num=2,
                           coll_num=column_num,
                           bolt=False)
        xls_set_cell_value(worksheet,
                           mah_result,
                           row_num=3,
                           coll_num=column_num,
                           bolt=False)
        xls_set_cell_value(worksheet,
                           mwh_result,
                           row_num=4,
                           coll_num=column_num,
                           bolt=False)
    xls_expands_all_columns_width(worksheet)
    xls_save_workbook(workbook, 'output.xlsx')


if __name__ == '__main__':
    creator(battery_log_test_params)
